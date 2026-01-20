from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

archivo = 'Liga_MX.xlsx'

def crear_archivo_nuevo():
    """Crea un nuevo archivo Excel con la estructura de la Liga MX"""
    wb = Workbook()
    
    # Hoja de Tabla de Posiciones
    ws_tabla = wb.active
    ws_tabla.title = "Tabla de Posiciones"
    
    # Encabezados con estilo
    encabezados = ['Equipo', 'PJ', 'PG', 'PE', 'PP', 'GF', 'GC', 'DIF', 'Puntos']
    ws_tabla.append(encabezados)
    
    # Estilo para encabezados
    for col in range(1, len(encabezados) + 1):
        celda = ws_tabla.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFF", size=12)
        celda.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")
    
    # Equipos de la Liga MX (Apertura 2024)
    equipos = [
        'América', 'Guadalajara', 'Cruz Azul', 'Pumas UNAM',
        'Tigres UANL', 'Monterrey', 'Santos Laguna', 'Toluca',
        'León', 'Atlas', 'Pachuca', 'Tijuana',
        'Necaxa', 'Querétaro', 'Mazatlán', 'Puebla',
        'Juárez', 'San Luis'
    ]
    
    for equipo in equipos:
        ws_tabla.append([equipo, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Ajustar ancho de columnas
    ws_tabla.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_tabla.column_dimensions[col].width = 8
    
    # Hoja de Resultados por Jornada
    ws_jornadas = wb.create_sheet("Resultados")
    ws_jornadas.append(['Jornada', 'Equipo Local', 'Goles Local', 'Goles Visitante', 'Equipo Visitante'])
    
    # Estilo para encabezados de jornadas
    for col in range(1, 6):
        celda = ws_jornadas.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")
    
    ws_jornadas.column_dimensions['A'].width = 10
    ws_jornadas.column_dimensions['B'].width = 20
    ws_jornadas.column_dimensions['C'].width = 12
    ws_jornadas.column_dimensions['D'].width = 15
    ws_jornadas.column_dimensions['E'].width = 20
    
    wb.save(archivo)
    print(f"✅ Archivo '{archivo}' creado exitosamente")
    print(f"📊 {len(equipos)} equipos registrados")
    return wb

def obtener_equipos():
    """Obtiene la lista de equipos del archivo"""
    wb = load_workbook(archivo)
    ws = wb['Tabla de Posiciones']
    equipos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0]:
            equipos.append(fila[0])
    return equipos

def agregar_equipo(nombre_equipo):# agregar_equipo('Atlante')
    """Agrega un nuevo equipo a la tabla"""
    try:
        wb = load_workbook(archivo)
        ws = wb['Tabla de Posiciones']
        
        # Verificar si el equipo ya existe
        equipos_existentes = [fila[0] for fila in ws.iter_rows(min_row=2, values_only=True) if fila[0]]
        
        if nombre_equipo in equipos_existentes:
            print(f"⚠️  El equipo '{nombre_equipo}' ya existe en la tabla")
            return False
        
        # Agregar el nuevo equipo
        ws.append([nombre_equipo, 0, 0, 0, 0, 0, 0, 0, 0])
        wb.save(archivo)
        print(f"➕ Equipo '{nombre_equipo}' agregado exitosamente")
        return True
        
    except Exception as e:
        print(f"❌ Error al agregar equipo: {e}")
        return False

def registrar_resultado(jornada, equipo_local, goles_local, goles_visitante, equipo_visitante):
    """Registra el resultado de un partido y actualiza la tabla"""
    try:
        wb = load_workbook(archivo)
        ws_tabla = wb['Tabla de Posiciones']
        ws_jornadas = wb['Resultados']
        
        # Validar que los equipos existan
        equipos = obtener_equipos()
        if equipo_local not in equipos:
            print(f"❌ El equipo '{equipo_local}' no existe")
            return False
        if equipo_visitante not in equipos:
            print(f"❌ El equipo '{equipo_visitante}' no existe")
            return False
        
        # Validar goles
        if not isinstance(goles_local, int) or not isinstance(goles_visitante, int):
            print("❌ Los goles deben ser números enteros")
            return False
        if goles_local < 0 or goles_visitante < 0:
            print("❌ Los goles no pueden ser negativos")
            return False
        
        # Determinar resultado
        if goles_local > goles_visitante:
            resultado_local = "Victoria"
            resultado_visitante = "Derrota"
            puntos_local = 3
            puntos_visitante = 0
        elif goles_local < goles_visitante:
            resultado_local = "Derrota"
            resultado_visitante = "Victoria"
            puntos_local = 0
            puntos_visitante = 3
        else:
            resultado_local = "Empate"
            resultado_visitante = "Empate"
            puntos_local = 1
            puntos_visitante = 1
        
        # Actualizar estadísticas de ambos equipos
        for fila in ws_tabla.iter_rows(min_row=2):
            equipo = fila[0].value
            
            if equipo == equipo_local:
                fila[1].value += 1  # PJ (Partidos Jugados)
                fila[5].value += goles_local  # GF (Goles a Favor)
                fila[6].value += goles_visitante  # GC (Goles en Contra)
                fila[7].value = fila[5].value - fila[6].value  # DIF (Diferencia)
                fila[8].value += puntos_local  # Puntos
                
                if resultado_local == "Victoria":
                    fila[2].value += 1  # PG (Partidos Ganados)
                elif resultado_local == "Empate":
                    fila[3].value += 1  # PE (Partidos Empatados)
                else:
                    fila[4].value += 1  # PP (Partidos Perdidos)
            
            elif equipo == equipo_visitante:
                fila[1].value += 1  # PJ
                fila[5].value += goles_visitante  # GF
                fila[6].value += goles_local  # GC
                fila[7].value = fila[5].value - fila[6].value  # DIF
                fila[8].value += puntos_visitante  # Puntos
                
                if resultado_visitante == "Victoria":
                    fila[2].value += 1  # PG
                elif resultado_visitante == "Empate":
                    fila[3].value += 1  # PE
                else:
                    fila[4].value += 1  # PP
        
        # Registrar en hoja de resultados
        ws_jornadas.append([jornada, equipo_local, goles_local, goles_visitante, equipo_visitante])
        
        # Ordenar tabla por puntos (de mayor a menor)
        datos = []
        for fila in ws_tabla.iter_rows(min_row=2, values_only=True):
            if fila[0]:
                datos.append(list(fila))
        
        # Ordenar por: Puntos (desc), Diferencia de goles (desc), Goles a favor (desc)
        datos.sort(key=lambda x: (x[8], x[7], x[5]), reverse=True)
        
        # Limpiar datos antiguos
        ws_tabla.delete_rows(2, ws_tabla.max_row)
        
        # Escribir datos ordenados
        for dato in datos:
            ws_tabla.append(dato)
        
        wb.save(archivo)
        
        print(f"\n⚽ JORNADA {jornada}")
        print(f"{'='*50}")
        print(f"{equipo_local} {goles_local} - {goles_visitante} {equipo_visitante}")
        print(f"{'='*50}")
        print(f"🏠 {equipo_local}: {resultado_local} (+{puntos_local} puntos)")
        print(f"✈️  {equipo_visitante}: {resultado_visitante} (+{puntos_visitante} puntos)")
        
        return True
        
    except Exception as e:
        print(f"❌ Error al registrar resultado: {e}")
        return False

def mostrar_tabla():
    """Muestra la tabla de posiciones en consola"""
    try:
        wb = load_workbook(archivo)
        ws = wb['Tabla de Posiciones']
        
        print(f"\n{'='*90}")
        print(f"{'TABLA DE POSICIONES - LIGA MX':^90}")
        print(f"{'='*90}")
        print(f"{'Pos':<5} {'Equipo':<20} {'PJ':<5} {'PG':<5} {'PE':<5} {'PP':<5} {'GF':<5} {'GC':<5} {'DIF':<6} {'Pts':<5}")
        print(f"{'-'*90}")
        
        pos = 1
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if fila[0]:
                print(f"{pos:<5} {fila[0]:<20} {fila[1]:<5} {fila[2]:<5} {fila[3]:<5} {fila[4]:<5} {fila[5]:<5} {fila[6]:<5} {fila[7]:<6} {fila[8]:<5}")
                pos += 1
        
        print(f"{'='*90}\n")
        
    except Exception as e:
        print(f"❌ Error al mostrar tabla: {e}")

# ============================================
# PROGRAMA PRINCIPAL
# ============================================

def main():
    """Función principal del programa"""
    
    # Crear archivo si no existe
    if not os.path.exists(archivo):
        print("📁 Creando nuevo archivo de Liga MX...")
        crear_archivo_nuevo()
    else:
        print(f"✅ Archivo '{archivo}' encontrado")
    
    # Ejemplo de uso: Registrar resultados de la Jornada 1
    print("\n" + "="*50)
    print("REGISTRANDO RESULTADOS - JORNADA 1")
    print("="*50)
    
    # Algunos partidos de ejemplo
    registrar_resultado(4, 'Juárez', 2, 1, 'cruz azul')
    registrar_resultado(4, 'Querétaro', 1, 2, 'Pachuca')
    registrar_resultado(4, 'Atlas', 1, 0, 'Mazatlan')
    registrar_resultado(4, 'Monterrey', 0, 0, 'Tijuana')
    registrar_resultado(4, 'Pumas', 2, 0, 'Santos Laguna')
    registrar_resultado(4, 'puebla', 0, 1, 'toluca')
    registrar_resultado(4, 'America', 1, 3, 'Necaxa')
    registrar_resultado(4, 'León', 1, 1, 'Tigres')
    registrar_resultado(4, 'San Luis', 1, 2, 'Guadalajara')
    
    # Mostrar tabla actualizada
    mostrar_tabla()
    
    # Abrir archivo
    try:
        os.startfile(archivo)
        print(f"📂 Abriendo {archivo}...")
    except Exception as e:
        print(f"ℹ️  Abre manualmente: {os.path.abspath(archivo)}")

if __name__ == "__main__":
    main()