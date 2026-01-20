# ============================================
# IMPORTACIONES
# ============================================
# Estas deben ir SIEMPRE al inicio del archivo (PEP 8 - Python Enhancement Proposal)

from openpyxl import load_workbook, Workbook
# - Workbook: Para CREAR archivos Excel nuevos desde cero
# - load_workbook: Para ABRIR archivos Excel existentes

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# Importamos clases de estilo para hacer el Excel más profesional:
# - Font: Cambiar tipo, tamaño, color de letra, negrita, etc.
# - PatternFill: Rellenar celdas con colores de fondo
# - Alignment: Alinear texto (centro, izquierda, derecha)
# - Border, Side: Agregar bordes a las celdas (no usado en este código pero importado por si acaso)

import os
# Módulo del sistema operativo para:
# - Verificar si archivos existen (os.path.exists)
# - Abrir archivos automáticamente (os.startfile)
# - Obtener rutas absolutas (os.path.abspath)

# ============================================
# CONSTANTES GLOBALES
# ============================================
# Se define al inicio para evitar escribir el nombre varias veces
# Si necesitas cambiar el nombre del archivo, solo lo cambias aquí
archivo = 'Liga_MX.xlsx'

# ============================================
# FUNCIÓN 1: CREAR ARCHIVO NUEVO
# ============================================
def crear_archivo_nuevo():
    """Crea un nuevo archivo Excel con la estructura de la Liga MX"""
    # Docstring: Explica qué hace la función (buena práctica)
    
    # Paso 1: Crear un libro de trabajo nuevo (archivo Excel vacío)
    wb = Workbook()
    # wb = workbook (libro de trabajo)
    # Es como abrir Excel y tener un archivo nuevo sin guardar
    
    # Paso 2: Obtener la hoja activa (la primera hoja que viene por defecto)
    ws_tabla = wb.active
    # ws = worksheet (hoja de cálculo)
    # Todo Workbook nuevo viene con UNA hoja por defecto
    
    # Paso 3: Renombrar la hoja
    ws_tabla.title = "Tabla de Posiciones"
    # Sin esto, se llamaría "Sheet" o "Hoja1"
    
    # Paso 4: Definir los encabezados de la tabla
    encabezados = ['Equipo', 'PJ', 'PE', 'PP', 'GF', 'GC', 'DIF', 'Puntos']
    # PJ = Partidos Jugados, PG = Partidos Ganados, PE = Partidos Empatados
    # PP = Partidos Perdidos, GF = Goles a Favor, GC = Goles en Contra
    # DIF = Diferencia de goles, Puntos = Puntuación total
    # Se usa una LISTA porque es más fácil de manejar que escribir celda por celda
    
    # Paso 5: Agregar los encabezados a la primera fila
    ws_tabla.append(encabezados)
    # append() agrega una fila completa al final de la hoja
    # Como la hoja está vacía, se agrega en la fila 1
    
    # Paso 6: Aplicar ESTILO a los encabezados (hacerlos bonitos)
    for col in range(1, len(encabezados) + 1):
        # range(1, 10) genera números del 1 al 9 (no incluye el 10)
        # ¿Por qué desde 1? Porque Excel cuenta columnas desde 1 (A=1, B=2, C=3...)
        # len(encabezados) + 1 = 9 + 1 = 10, entonces va de 1 a 9
        
        celda = ws_tabla.cell(row=1, column=col)
        # Obtener la celda específica de la fila 1, columna actual
        # Ejemplo: col=1 → celda A1, col=2 → celda B1
        
        celda.font = Font(bold=True, color="FFFFFF", size=12)
        # Aplicar formato de letra:
        # - bold=True: Texto en negrita
        # - color="FFFFFF": Color blanco (formato hexadecimal)
        # - size=12: Tamaño de letra 12 puntos
        
        celda.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        # Rellenar el fondo de la celda:
        # - start_color y end_color iguales = color sólido
        # - "1F4E78" = Azul oscuro profesional
        # - fill_type="solid" = Relleno sólido (no degradado)
        
        celda.alignment = Alignment(horizontal="center", vertical="center")
        # Alinear el texto:
        # - horizontal="center": Centrado horizontalmente
        # - vertical="center": Centrado verticalmente
        # Hace que los encabezados se vean más profesionales
    
    # Paso 7: Definir los equipos de la Liga MX
    equipos = [
        'América', 'Guadalajara', 'Cruz Azul', 'Pumas UNAM',
        'Tigres UANL', 'Monterrey', 'Santos Laguna', 'Toluca',
        'León', 'Atlas', 'Pachuca', 'Tijuana',
        'Necaxa', 'Querétaro', 'Mazatlán', 'Puebla',
        'Juárez', 'San Luis'
    ]
    # Son 18 equipos actuales de la Liga MX (Apertura 2024)
    # Se usa una lista para poder agregarlos fácilmente con un loop
    
    # Paso 8: Agregar cada equipo con estadísticas iniciales en 0
    for equipo in equipos:
        # Por cada equipo en la lista, agregar una fila
        ws_tabla.append([equipo, 0, 0, 0, 0, 0, 0, 0, 0])
        # Formato: [Equipo, PJ, PG, PE, PP, GF, GC, DIF, Puntos]
        # Todos empiezan en 0 porque no han jugado partidos
        # ¿Por qué 9 valores? Porque tenemos 9 columnas (encabezados)
    
    # Paso 9: Ajustar el ancho de las columnas para mejor visualización
    ws_tabla.column_dimensions['A'].width = 20
    # Columna A (Equipo) más ancha porque los nombres son largos
    
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_tabla.column_dimensions[col].width = 8
        # Las demás columnas son números, no necesitan tanto espacio
        # Se recorren todas de una vez para no repetir código
    
    # Paso 10: Crear una SEGUNDA hoja para los resultados
    ws_jornadas = wb.create_sheet("Resultados")
    # create_sheet() crea una nueva hoja en el mismo archivo
    # Ahora tenemos 2 hojas: "Tabla de Posiciones" y "Resultados"
    
    # Paso 11: Agregar encabezados a la hoja de resultados
    ws_jornadas.append(['Jornada', 'Equipo Local', 'Goles Local', 'Goles Visitante', 'Equipo Visitante'])
    # Esta hoja llevará el historial de todos los partidos jugados
    
    # Paso 12: Aplicar estilo a los encabezados de la hoja de resultados
    for col in range(1, 6):  # 5 columnas en esta hoja
        celda = ws_jornadas.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")
        # Mismo estilo que la tabla principal para consistencia visual
    
    # Paso 13: Ajustar anchos de columnas de la hoja de resultados
    ws_jornadas.column_dimensions['A'].width = 10   # Jornada (número pequeño)
    ws_jornadas.column_dimensions['B'].width = 20   # Equipo Local (nombre largo)
    ws_jornadas.column_dimensions['C'].width = 12   # Goles Local
    ws_jornadas.column_dimensions['D'].width = 15   # Goles Visitante
    ws_jornadas.column_dimensions['E'].width = 20   # Equipo Visitante (nombre largo)
    
    # Paso 14: GUARDAR el archivo
    wb.save(archivo)
    # Sin este paso, todo lo anterior solo existe en memoria (RAM)
    # save() escribe físicamente el archivo en el disco duro
    
    # Paso 15: Mostrar mensajes de confirmación
    print(f"✅ Archivo '{archivo}' creado exitosamente")
    print(f"📊 {len(equipos)} equipos registrados")
    # f-strings (f"...{variable}...") permiten insertar variables en texto
    # len(equipos) cuenta cuántos elementos tiene la lista (18)
    
    # Paso 16: Retornar el workbook por si se necesita usar después
    return wb
    # Aunque no es obligatorio, es buena práctica retornar algo útil

# ============================================
# FUNCIÓN 2: OBTENER EQUIPOS
# ============================================
def obtener_equipos():
    """Obtiene la lista de equipos del archivo"""
    # Esta función es AUXILIAR, la usarán otras funciones
    
    wb = load_workbook(archivo)
    # Abrimos el archivo existente (NO creamos uno nuevo)
    # load_workbook = "cargar libro de trabajo"
    
    ws = wb['Tabla de Posiciones']
    # Accedemos a la hoja específica por su nombre
    # Es como hacer clic en la pestaña "Tabla de Posiciones" en Excel
    
    equipos = []
    # Creamos una lista vacía para ir guardando los nombres
    
    for fila in ws.iter_rows(min_row=2, values_only=True):
        # iter_rows = iterar (recorrer) filas
        # min_row=2: Empezar desde la fila 2 (saltarse encabezados en fila 1)
        # values_only=True: Solo obtener valores, no objetos Cell
        #   Sin values_only: fila = [<Cell 'A2'>, <Cell 'B2'>, ...]
        #   Con values_only: fila = ['América', 0, 0, 0, ...]
        
        if fila[0]:
            # Verificar que la primera celda (nombre del equipo) no esté vacía
            # fila[0] = primer elemento de la fila (columna A)
            # Esto evita agregar filas vacías si existen
            
            equipos.append(fila[0])
            # Agregar solo el nombre del equipo a la lista
    
    return equipos
    # Retornar la lista completa de equipos
    # Ejemplo de retorno: ['América', 'Guadalajara', 'Cruz Azul', ...]

# ============================================
# FUNCIÓN 3: AGREGAR EQUIPO
# ============================================
def agregar_equipo(nombre_equipo):
    """Agrega un nuevo equipo a la tabla"""
    # Parámetro: nombre_equipo = el nombre que queremos agregar
    
    try:
        # try-except: Bloque para MANEJAR ERRORES
        # Si algo sale mal dentro del try, salta al except
        # Esto evita que el programa se detenga abruptamente
        
        wb = load_workbook(archivo)
        # Abrir el archivo existente
        
        ws = wb['Tabla de Posiciones']
        # Acceder a la hoja de tabla
        
        # Paso 1: Obtener todos los equipos que ya existen
        equipos_existentes = [fila[0] for fila in ws.iter_rows(min_row=2, values_only=True) if fila[0]]
        # Esto es una LIST COMPREHENSION (forma compacta de crear listas)
        # Equivalente a:
        # equipos_existentes = []
        # for fila in ws.iter_rows(min_row=2, values_only=True):
        #     if fila[0]:
        #         equipos_existentes.append(fila[0])
        
        # Paso 2: Verificar si el equipo YA existe
        if nombre_equipo in equipos_existentes:
            # El operador 'in' verifica si un elemento está en una lista
            print(f"⚠️  El equipo '{nombre_equipo}' ya existe en la tabla")
            return False
            # return False = indicar que NO se agregó el equipo
            # Al hacer return, la función termina aquí (no ejecuta lo que sigue)
        
        # Paso 3: Si llegamos aquí, el equipo NO existe, entonces agregarlo
        ws.append([nombre_equipo, 0, 0, 0, 0, 0, 0, 0, 0])
        # Agregar nueva fila con estadísticas en 0
        
        wb.save(archivo)
        # IMPORTANTE: Guardar los cambios
        
        print(f"➕ Equipo '{nombre_equipo}' agregado exitosamente")
        return True
        # return True = indicar que SÍ se agregó correctamente
        
    except Exception as e:
        # except captura CUALQUIER error que ocurra en el try
        # Exception es la clase base de todos los errores
        # 'as e' guarda el error en la variable 'e'
        
        print(f"❌ Error al agregar equipo: {e}")
        # Mostrar qué error ocurrió
        # Ejemplo de error: "FileNotFoundError: Liga_MX.xlsx not found"
        
        return False
        # Indicar que algo salió mal

# ============================================
# FUNCIÓN 4: REGISTRAR RESULTADO (LA MÁS IMPORTANTE)
# ============================================
def registrar_resultado(jornada, equipo_local, goles_local, goles_visitante, equipo_visitante):
    """Registra el resultado de un partido y actualiza la tabla"""
    # Parámetros:
    # - jornada: número de jornada (1-17)
    # - equipo_local: nombre del equipo que juega en casa
    # - goles_local: goles que anotó el equipo local
    # - goles_visitante: goles que anotó el equipo visitante
    # - equipo_visitante: nombre del equipo que juega de visita
    
    try:
        # === PASO 1: ABRIR EL ARCHIVO ===
        wb = load_workbook(archivo)
        ws_tabla = wb['Tabla de Posiciones']  # Hoja de estadísticas
        ws_jornadas = wb['Resultados']        # Hoja de historial
        
        # === PASO 2: VALIDAR QUE LOS EQUIPOS EXISTAN ===
        equipos = obtener_equipos()
        # Llamamos a la función que creamos antes
        # equipos es una lista: ['América', 'Guadalajara', ...]
        
        if equipo_local not in equipos:
            # Verificar que el equipo local esté registrado
            print(f"❌ El equipo '{equipo_local}' no existe")
            return False  # Terminar la función si no existe
        
        if equipo_visitante not in equipos:
            # Verificar que el equipo visitante esté registrado
            print(f"❌ El equipo '{equipo_visitante}' no existe")
            return False
        
        # === PASO 3: VALIDAR QUE LOS GOLES SEAN NÚMEROS VÁLIDOS ===
        if not isinstance(goles_local, int) or not isinstance(goles_visitante, int):
            # isinstance(variable, tipo) verifica si una variable es de cierto tipo
            # Ejemplo: isinstance(5, int) = True
            # Ejemplo: isinstance("5", int) = False
            # Esto evita que alguien ponga "tres" en lugar de 3
            
            print("❌ Los goles deben ser números enteros")
            return False
        
        if goles_local < 0 or goles_visitante < 0:
            # No pueden haber goles negativos
            print("❌ Los goles no pueden ser negativos")
            return False
        
        # === PASO 4: DETERMINAR EL RESULTADO (QUIÉN GANÓ) ===
        if goles_local > goles_visitante:
            # Si el local anotó más goles → GANÓ el local
            resultado_local = "Victoria"
            resultado_visitante = "Derrota"
            puntos_local = 3       # Victoria = 3 puntos
            puntos_visitante = 0   # Derrota = 0 puntos
            
        elif goles_local < goles_visitante:
            # Si el visitante anotó más goles → GANÓ el visitante
            resultado_local = "Derrota"
            resultado_visitante = "Victoria"
            puntos_local = 0
            puntos_visitante = 3
            
        else:
            # Si anotaron lo mismo → EMPATE
            resultado_local = "Empate"
            resultado_visitante = "Empate"
            puntos_local = 1       # Empate = 1 punto para cada uno
            puntos_visitante = 1
        
        # === PASO 5: ACTUALIZAR ESTADÍSTICAS DE AMBOS EQUIPOS ===
        for fila in ws_tabla.iter_rows(min_row=2):
            # Recorrer TODAS las filas de equipos (desde fila 2)
            # IMPORTANTE: Aquí NO usamos values_only=True
            # ¿Por qué? Porque necesitamos MODIFICAR las celdas
            # values_only solo nos daría los valores, no las celdas editables
            
            equipo = fila[0].value
            # fila[0] = objeto Cell de la columna A
            # fila[0].value = el valor dentro de esa celda (nombre del equipo)
            
            # --- ACTUALIZAR EQUIPO LOCAL ---
            if equipo == equipo_local:
                # Si encontramos la fila del equipo local
                
                fila[1].value += 1  
                # fila[1] = columna B = PJ (Partidos Jugados)
                # += 1 significa: sumar 1 al valor actual
                # Equivalente a: fila[1].value = fila[1].value + 1
                
                fila[5].value += goles_local
                # fila[5] = columna F = GF (Goles a Favor)
                # Sumar los goles que anotó
                
                fila[6].value += goles_visitante
                # fila[6] = columna G = GC (Goles en Contra)
                # Sumar los goles que le anotaron
                
                fila[7].value = fila[5].value - fila[6].value
                # fila[7] = columna H = DIF (Diferencia de goles)
                # DIF = Goles a favor - Goles en contra
                # Se RECALCULA cada vez (no se suma)
                
                fila[8].value += puntos_local
                # fila[8] = columna I = Puntos
                # Sumar los puntos obtenidos (0, 1, o 3)
                
                # Actualizar contadores de victorias/empates/derrotas
                if resultado_local == "Victoria":
                    fila[2].value += 1  # PG (Partidos Ganados)
                elif resultado_local == "Empate":
                    fila[3].value += 1  # PE (Partidos Empatados)
                else:
                    fila[4].value += 1  # PP (Partidos Perdidos)
            
            # --- ACTUALIZAR EQUIPO VISITANTE ---
            elif equipo == equipo_visitante:
                # Si encontramos la fila del equipo visitante
                # La lógica es EXACTAMENTE igual que el local
                # pero con los valores invertidos
                
                fila[1].value += 1  # PJ
                fila[5].value += goles_visitante  # Sus goles a favor
                fila[6].value += goles_local       # Sus goles en contra
                fila[7].value = fila[5].value - fila[6].value  # DIF
                fila[8].value += puntos_visitante  # Puntos
                
                if resultado_visitante == "Victoria":
                    fila[2].value += 1  # PG
                elif resultado_visitante == "Empate":
                    fila[3].value += 1  # PE
                else:
                    fila[4].value += 1  # PP
        
        # === PASO 6: REGISTRAR EL PARTIDO EN LA HOJA DE RESULTADOS ===
        ws_jornadas.append([jornada, equipo_local, goles_local, goles_visitante, equipo_visitante])
        # Agregar una nueva fila con toda la información del partido
        # Esto crea un HISTORIAL de todos los partidos jugados
        
        # === PASO 7: ORDENAR LA TABLA POR PUNTOS ===
        # Primero, leer todos los datos actuales
        datos = []
        for fila in ws_tabla.iter_rows(min_row=2, values_only=True):
            # Ahora SÍ usamos values_only porque solo queremos leer
            if fila[0]:  # Si la fila tiene un equipo
                datos.append(list(fila))
                # list(fila) convierte la tupla en lista (para poder modificarla)
        
        # Ordenar la lista de datos
        datos.sort(key=lambda x: (x[8], x[7], x[5]), reverse=True)
        # Esta línea es COMPLEJA, desglosémosla:
        # 
        # .sort() = ordenar la lista
        # key= especifica CÓMO ordenar
        # lambda x: ... = función anónima (función temporal)
        #   - x representa cada equipo (una lista de 9 valores)
        #   - (x[8], x[7], x[5]) = tupla con 3 criterios de ordenamiento
        #     1. x[8] = Puntos (más importante)
        #     2. x[7] = Diferencia de goles (si empatan en puntos)
        #     3. x[5] = Goles a favor (si empatan en puntos y diferencia)
        # reverse=True = de mayor a menor (descendente)
        #
        # Ejemplo:
        # Equipo A: 15 puntos, +5 diferencia, 20 goles
        # Equipo B: 15 puntos, +5 diferencia, 18 goles
        # Equipo C: 15 puntos, +3 diferencia, 22 goles
        # Orden: A, B, C (mismo puntos y DIF, pero A tiene más goles)
        
        # Limpiar la tabla actual
        ws_tabla.delete_rows(2, ws_tabla.max_row)
        # delete_rows(inicio, cuántas)
        # Borramos desde fila 2 hasta el final
        # ¿Por qué? Para escribir los datos ordenados desde cero
        
        # Escribir los datos ordenados
        for dato in datos:
            ws_tabla.append(dato)
            # Ahora los equipos están ordenados por posición en la tabla
        
        # === PASO 8: GUARDAR LOS CAMBIOS ===
        wb.save(archivo)
        # CRÍTICO: Sin esto, todos los cambios se pierden
        
        # === PASO 9: MOSTRAR RESUMEN DEL PARTIDO ===
        print(f"\n⚽ JORNADA {jornada}")
        print(f"{'='*50}")
        # '='*50 crea una línea de 50 signos de igual
        # Es para hacer una separación visual bonita
        
        print(f"{equipo_local} {goles_local} - {goles_visitante} {equipo_visitante}")
        print(f"{'='*50}")
        print(f"🏠 {equipo_local}: {resultado_local} (+{puntos_local} puntos)")
        print(f"✈️  {equipo_visitante}: {resultado_visitante} (+{puntos_visitante} puntos)")
        # Los emojis hacen la salida más visual y amigable
        
        return True
        # Indicar que todo salió bien
        
    except Exception as e:
        # Si CUALQUIER cosa sale mal en todo el proceso
        print(f"❌ Error al registrar resultado: {e}")
        # Mostrar qué error ocurrió
        return False
        # Indicar que hubo un problema

# ============================================
# FUNCIÓN 5: MOSTRAR TABLA EN CONSOLA
# ============================================
def mostrar_tabla():
    """Muestra la tabla de posiciones en consola"""
    # Esta función es para VER la tabla sin abrir Excel
    
    try:
        wb = load_workbook(archivo)
        ws = wb['Tabla de Posiciones']
        
        # Encabezado decorativo
        print(f"\n{'='*90}")
        # {'TEXTO':^90} = centrar texto en un espacio de 90 caracteres
        print(f"{'TABLA DE POSICIONES - LIGA MX':^90}")
        print(f"{'='*90}")
        
        # Encabezados de las columnas
        # {:<5} = alinear a la izquierda en 5 espacios
        # {:^90} = alinear al centro en 90 espacios
        print(f"{'Pos':<5} {'Equipo':<20} {'PJ':<5} {'PG':<5} {'PE':<5} {'PP':<5} {'GF':<5} {'GC':<5} {'DIF':<6} {'Pts':<5}")
        print(f"{'-'*90}")
        # Línea separadora
        
        pos = 1  # Contador de posición
        for fila in ws.iter_rows(min_row=2, values_only=True):
            # Recorrer todas las filas de equipos
            if fila[0]:  # Si hay un equipo en esa fila
                # fila[0] = Equipo, fila[1] = PJ, fila[2] = PG, etc.
                print(f"{pos:<5} {fila[0]:<20} {fila[1]:<5} {fila[2]:<5} {fila[3]:<5} {fila[4]:<5} {fila[5]:<5} {fila[6]:<5} {fila[7]:<6} {fila[8]:<5}")
                pos += 1  # Incrementar posición para el siguiente equipo
        
        print(f"{'='*90}\n")
        # Línea final de cierre
        
    except Exception as e:
        print(f"❌ Error al mostrar tabla: {e}")

# ============================================
# FUNCIÓN 6: PROGRAMA PRINCIPAL (main)
# ============================================
def main():
    """Función principal del programa"""
    # Esta es la función que ORQUESTA todo
    # Es la que se ejecuta cuando corres el programa
    
    # === PASO 1: VERIFICAR SI EL ARCHIVO EXISTE ===
    if not os.path.exists(archivo):
        # os.path.exists() verifica si un archivo existe en el disco
        # not invierte el resultado (True → False, False → True)
        # Si el archivo NO existe, entrar aquí
        
        print("📁 Creando nuevo archivo de Liga MX...")
        crear_archivo_nuevo()
        # Llamar a la función que crea el archivo desde cero
        
    else:
        # Si el archivo SÍ existe
        print(f"✅ Archivo '{archivo}' encontrado")
    
    # === PASO 2: REGISTRAR RESULTADOS DE EJEMPLO ===
    print("\n" + "="*50)
    # "\n" = salto de línea (nueva línea)
    # + concatena strings
    
    print("REGISTRANDO RESULTADOS - JORNADA 1")
    print("="*50)
    
    # Algunos partidos de ejemplo de la Jornada 1
    # Estos son solo para demostración
    # En un uso real, cambiarías estos datos cada vez
    
    registrar_resultado(1, 'América', 2, 1, 'Guadalajara')
    # Jornada 1: América 2-1 Guadalajara
    # América gana: +3 puntos
    # Guadalajara pierde: +0 puntos
    
    registrar_resultado(1, 'Cruz Azul', 3, 0, 'Pumas UNAM')
    # Cruz Azul 3-0 Pumas
    
    registrar_resultado(1, 'Tigres UANL', 1, 1, 'Monterrey')
    # Tigres 1-1 Monterrey (empate)
    # Ambos equipos: +1 punto
    
    registrar_resultado(1, 'Santos Laguna', 0, 2, 'Toluca')
    # Santos 0-2 Toluca
    # Toluca gana de visitante: +3 puntos
    
    # === PASO 3: MOSTRAR LA TABLA ACTUALIZADA ===
    mostrar_tabla()
    # Llamar a la función que imprime la tabla en consola
    
    # === PASO 4: ABRIR EL ARCHIVO AUTOMÁTICAMENTE ===
    try:
        os.startfile(archivo)
        # os.startfile() abre un archivo con su programa predeterminado
        # En Windows, abre archivos .xlsx con Excel
        # Solo funciona en Windows
        
        print(f"📂 Abriendo {archivo}...")
        
    except Exception as e:
        # Si no puede abrir (por ejemplo, en Mac o Linux)
        print(f"ℹ️  Abre manualmente: {os.path.abspath(archivo)}")
        # os.path.abspath() da la ruta completa del archivo
        # Ejemplo: C:\Users\usuario\Liga_MX.xlsx

# ============================================
# PUNTO DE ENTRADA DEL PROGRAMA
# ============================================
if __name__ == "__main__":
    # Esta línea es MUY IMPORTANTE en Python
    # 
    # __name__ es una variable especial que Python crea automáticamente
    # 
    # Cuando ejecutas un archivo directamente:
    #   python excel-03.py
    #   → __name__ = "__main__"
    # 
    # Cuando importas un archivo desde otro:
    #   import excel-03
    #   → __name__ = "excel-03"
    # 
    # ¿Para qué sirve esto?
    # - Si ejecutas el archivo directamente → corre main()
    # - Si lo importas → NO corre main() automáticamente
    #   (pero puedes usar las funciones definidas)
    # 
    # Esto permite que el mismo archivo funcione como:
    #   1. Programa ejecutable (si lo corres directamente)
    #   2. Módulo/librería (si lo importas en otro archivo)
    
    main()
    # Llamar a la función principal
    # Aquí empieza la ejecución real del programa

# ============================================
# ORDEN DE EJECUCIÓN AL CORRER EL PROGRAMA
# ============================================
# 1. Python lee TODO el archivo de arriba hacia abajo
# 2. Importa las librerías (openpyxl, os)
# 3. Define la variable 'archivo'
# 4. Define las funciones (pero NO las ejecuta aún)
# 5. Llega a if __name__ == "__main__":
# 6. Como se cumple, ejecuta main()
# 7. main() verifica si existe el archivo
# 8. main() registra resultados
# 9. main() muestra la tabla
# 10. main() intenta abrir el archivo
# 11. Programa termina

# ============================================
# POR QUÉ ESTE ORDEN EN EL CÓDIGO
# ============================================
# 
# 1. IMPORTACIONES AL INICIO
#    - Es convención de Python (PEP 8)
#    - Fácil ver qué dependencias usa el programa
# 
# 2. CONSTANTES DESPUÉS DE IMPORTACIONES
#    - Variables que no cambian durante la ejecución
# 
# 3. FUNCIONES EN ORDEN LÓGICO
#    - crear_archivo_nuevo() → primera vez
#    - obtener_equipos() → auxiliar (usada por otras)
#    - agregar_equipo() → operación simple
#    - registrar_resultado() → operación compleja (usa otras)
#    - mostrar_tabla() → visualización
#    - main() → orquestadora (usa todas las anteriores)
# 
# 4. if __name__ == "__main__": AL FINAL
#    - Todas las funciones ya están definidas
#    - Solo se ejecuta si corres el archivo directamente
# 
# Este orden hace el código:
# - Fácil de leer (de arriba hacia abajo)
# - Fácil de mantener (cada función hace UNA cosa)
# - Reutilizable (puedes importar funciones en otros archivos)
# - Profesional (sigue estándares de Python)